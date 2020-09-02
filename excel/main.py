import openpyxl

xlsl = "D:\code\python\excel\main.xlsx"

def main():
    wb = openpyxl.load_workbook(xlsl)
    print(wb.sheetnames)
    ws = wb.active
    print(ws)
    print(ws['A1'].value)
    for i in range(1,5,1):
        print(ws.calculate_dimension())
        print(ws.cell(i,1).value)
        print(ws.cell(i, 2).value)
        print(ws.cell(i, 3).value)


if __name__ == '__main__':
    main()